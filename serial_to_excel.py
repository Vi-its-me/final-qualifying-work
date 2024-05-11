import serial
import time
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Border, Side

got_begin = False
workbook = Workbook()
worksheet = workbook.active
print_condition_flag = False
serial_object = serial.Serial()
port_number = 5
receiving_terminated = False
begin_found_flag = False
# Исключение для вызова в случае, если не получили
class Did_not_begin(Exception):     # "стартовую" строку
    def __init__(self, text):
        self.txt = text
# Функция для печати в таблицу, начиная с первой её строки
def insert_row_to_spreadsheet(row_number, list):
    for index, element in enumerate(list, 1):
        if isinstance(element, str) and\
            worksheet.column_dimensions\
            [get_column_letter(index)].width <\
            len(element) + 1:

            worksheet.column_dimensions\
            [get_column_letter(index)].width = \
            len(element) + 1

        worksheet.cell(row=row_number,
        column=index).value = element
# Функция для поиска первой пустой ячейки строки
# в столбце, переданном параметром
def find_empty_row(column_number = 1):
    row_number = 1
    while row_number != 30:
        if worksheet[get_column_letter(column_number)
        + str(row_number)].value is None:
            return row_number
        row_number += 1
def find_average(column_number):
    count = total = 0
    for row_number in range(2, 12):
        if worksheet.cell(row=row_number,
        column=column_number).value is None:
            continue
        elif worksheet.cell(row=row_number,
        column=column_number).value == 404:
            worksheet.cell(row=row_number,
            column=column_number).value = "—"
            continue
        total += int(worksheet.cell(row=row_number,
        column=column_number).value)
        count += 1
    return total / count
def insert_list_horizontally_into_spreadsheet(
column_number, start_row, to_insert):
    for index, element in enumerate(to_insert, start_row):
        worksheet.cell(row=index,
        column=column_number).value = element
def determine_and_insert_distance(signal_amount_column,
distance_column):
    for current_row in range(2, 12):
        if int(worksheet.cell(row=current_row,
        column=signal_amount_column).value) >= -50:
            worksheet.cell(row=current_row,
            column=distance_column).value = (abs(int(
            worksheet.cell(row=current_row,
            column=signal_amount_column)
            .value)) + (-10.09)) / 43.893
        elif int(worksheet.cell(row=current_row,
        column=signal_amount_column).value) < -50:
            worksheet.cell(row=current_row,
            column=distance_column).value = abs((abs(int(
            worksheet.cell(row=current_row,
            column=signal_amount_column)
            .value)) + (-54.698)) / 0.457)
def merge_rows_and_insert(column_number, to_insert):
    row_number = 2
    worksheet.merge_cells(start_row=row_number,
    end_row=row_number + 9, start_column=column_number,
    end_column=column_number)
    worksheet.cell(row=row_number,
    column=column_number).value = to_insert
def convert_column_to_int(column_number, start_row):
    for row_number in range(start_row, start_row + 10):
        worksheet.cell(row=row_number,
        column=column_number).value = int(
        worksheet.cell(row=row_number,
        column=column_number).value)
def set_style_to_row(row_number, style):
    column_number = 1
    while worksheet.cell(row=row_number,
    column=column_number).value is not None:
        if isinstance(style, Border):
            worksheet.cell(row=row_number,
            column=column_number).border = style
        # Не используется
        elif isinstance(style, Side):
            worksheet.cell(row=row_number,
            column=column_number).side = style
        column_number += 1
# Попытка открыть порт, пока тот не будет открыт
while not serial_object.is_open:
    try:
        serial_object = serial.Serial(port="COM"+str(
        port_number), baudrate=115200, timeout=20)
    except serial.SerialException:
        if not print_condition_flag:
            print("Port serial is not opened", end="")
            print_condition_flag = True
        if print_condition_flag:
            print(".", end="")
        time.sleep(1)
print("Port opened")
# Заполнение первой строки называниями столбцов
print_list = ["№ замера", "Уровень сигнала(dBm)",
"Время прохождения сигнала(μs)",
"Расстояние от уровня сигнала(m)","Модальное значение(m)",
"Погрешность измерения(%)", "Фактическое расстояние(m)"]
insert_row_to_spreadsheet(1, print_list)
# Стилизация 1 строки таблицы
side_style = Side(border_style='thick', color='00000000')
set_style_to_row(1, Border(top=side_style,
bottom=side_style))
# Заполнение таблицы данными, полученными от NodeMCU
worksheet_row = 2
try:
    while True:
        serial_string = serial_object.readline().decode()
        if serial_string == "end\n": # Переносом строки
            if not begin_found_flag: # может стать "\r\n"
                raise Did_not_begin("begin not found")
            break
        elif begin_found_flag:
            worksheet_column = 1
            previous_iterator = 0
            iterator = 0
            if iterator:
                previous_iterator = iterator
            for char in serial_string:
                if char == "\t" or char == "\n":
                    worksheet.cell(row=worksheet_row,
                    column=worksheet_column
                    ).value = serial_string[
                    previous_iterator:iterator]

                    worksheet_column += 1
                    previous_iterator = iterator + 1
                iterator += 1
            worksheet_row += 1
        elif serial_string.find("begin\n") != -1:
            begin_found_flag = True
    # Приведение всех чисел, полученные от NodeMCU к int
    for counter in range(1, 4):
        convert_column_to_int(counter, 2)
    # Подсчет среднего значения в каждой строке из
    # первого и печать в столбце из второго параметра
    determine_and_insert_distance(2, 4)
    # Объединение всех строк в заданном столбце
    # и подсчет моды всех значений
    merge_rows_and_insert(5, "=MODE(ROUND(D3:D11, 2))")
except Did_not_begin as how:
    print(how)
except serial.SerialException:
    print(f"Nothing got from COM{port_number}")
else:
    actual_distance = float(input(
    "Enter actual distance in meters: "))
    while actual_distance <= 0.0:
        actual_distance = float(input(
        "Invalid input\nEnter distance in range (0, +∞):"))
    # В последнем столбце объединение всех строк и
    # помещение значения фактического расстояния
    merge_rows_and_insert(7, actual_distance)
    # Вычисление погрешности измерения для каждого измер-я
    start_index = 2
    list_of_deviations = []
    for index in range(10):
        list_of_deviations.append(
        f"=ABS(({get_column_letter(4)}{index + 2} / {
        get_column_letter(7)}{start_index} * 100) - 100)")
    insert_list_horizontally_into_spreadsheet(6,
    start_index, list_of_deviations)
    # Подсчет среднего значения в столбцах и печать
    insert_row_to_spreadsheet(find_empty_row(),
        ["Среднее арифметическое:",
        find_average(2),
        find_average(3),
        "=AVERAGE(D3:D11)"])
    # Подсчет медианных значений в столбцах и печать
    insert_row_to_spreadsheet(find_empty_row(),
        ["Медианное значение:",
        f"=MEDIAN({get_column_letter(2)}{start_index}:\
        {get_column_letter(2)}{start_index + 9})",
        f"=MEDIAN({get_column_letter(3)}{start_index}:\
        {get_column_letter(3)}{start_index + 9})",
        f"=MEDIAN({get_column_letter(4)}{start_index}:\
        {get_column_letter(4)}{start_index + 9})"])
    # Стилизация таблицы
    set_style_to_row(12,  Border(top=side_style))
    # Сохранение в файл, названным фактическим расстоянием
    workbook.save(str(actual_distance) + "m" + ".xlsx")
finally:
    serial_object.close()
